from flask import Blueprint, render_template, request, jsonify, send_file
from utils.decorators import login_required
from Model.laporanModel import LaporanModel 
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

laporan_admin = Blueprint("laporan_admin",__name__,template_folder="../templates")

@laporan_admin.route("/laporan",methods=['GET'])
@login_required
def init():
    return render_template("admin/laporan_admin.html")

@laporan_admin.route("/reports/properties/excel", methods=['POST'])
@login_required
def keuangan_export():
    data = request.get_json()
    
    return data

@laporan_admin.route("/reports/properties/excel", methods=["POST"])
@login_required
def data_export_excel():
    data = request.get_json()

    result = LaporanModel.get_all_data_properties_from_transaksi(data)

    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Properties"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "LAPORAN DATA PROPERTIES"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Spasi baris
    ws.append([])

    headers = [
        "No",
        "Nama Property",
        "Harga",
        "Jenis Property",
        "Tanggal",
        "Status Transaksi"
    ]

    ws.append(headers)
    header_row = ws.max_row

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = 22

    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:F{header_row}"
    
    for idx, item in enumerate(result, start=1):
        ws.append([
            idx,
            item.get("nama_properti", "-"),
            item.get("harga", "-"),
            item.get("tipe_properti", "-"),
            item.get("jadwal_transaksi", "-"),
            item.get("status_transaksi", "-")
        ])

    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return send_file(
        excel_buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="laporan_properties.xlsx"
    )
    
    
@laporan_admin.route("/reports/properties/pdf", methods=["POST"])
@login_required
def data_export_pdf():
    data = request.get_json()
    result = LaporanModel.get_all_data_properties_from_transaksi(data)

    pdf_buffer = BytesIO()
    c = canvas.Canvas(pdf_buffer, pagesize=A4)
    width, height = A4

    def draw_header():
        c.setFont("Helvetica-Bold", 14)
        c.drawString(2 * cm, height - 2 * cm, "LAPORAN DATA PROPERTIES")

        c.setFont("Helvetica", 9)
        c.drawRightString(
            width - 2 * cm,
            height - 2 * cm,
            "Dicetak: " + datetime.now().strftime("%d-%m-%Y %H:%M")
        )

        # Header tabel
        y = height - 3 * cm
        c.setFont("Helvetica-Bold", 10)

        c.drawString(2 * cm, y, "No")
        c.drawString(3 * cm, y, "Nama Property")
        c.drawString(7 * cm, y, "Harga")
        c.drawString(10 * cm, y, "Jenis Property")
        c.drawString(13 * cm, y, "Tanggal")
        c.drawString(17 * cm, y, "Status")

        c.setLineWidth(0.5)
        c.line(2 * cm, y - 0.3 * cm, width - 2 * cm, y - 0.3 * cm)

        return y - 0.8 * cm

    y = draw_header()
    c.setFont("Helvetica", 9)

    for idx, item in enumerate(result, start=1):
        if y < 2.5 * cm:
            c.showPage()
            y = draw_header()
            c.setFont("Helvetica", 9)

        c.drawString(2 * cm, y, str(idx))
        c.drawString(3 * cm, y, str(item.get("nama_properti", "-")))

        # Format harga
        harga = item.get("harga", 0)
        harga_text = f"Rp {int(harga):,}".replace(",", ".") if isinstance(harga, (int, float)) else str(harga)
        c.drawRightString(9.5 * cm, y, harga_text)

        c.drawString(10 * cm, y, str(item.get("tipe_properti", "-")))
        c.drawString(13 * cm, y, str(item.get("jadwal_transaksi", "-")))
        c.drawString(17 * cm, y, str(item.get("status_transaksi", "-")))

        y -= 0.6 * cm

    # Footer page number
    page_count = c.getPageNumber()
    for page in range(1, page_count + 1):
        c.showPage()

    c.save()
    pdf_buffer.seek(0)

    return send_file(
        pdf_buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name="laporan_properties.pdf"
    )